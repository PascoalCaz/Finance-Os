import io
import os
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

class ReportService:
    """
    Serviço para geração de relatórios PDF e Excel.
    """

    @staticmethod
    def generate_pdf_report(transactions, summary):
        """
        Gera um relatório PDF formatado.
        Retorna: BytesIO object contendo o PDF.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        # Estilo para Título
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1, # Center
            textColor=colors.HexColor("#1e293b")
        )

        elements.append(Paragraph(f"Relatório Financeiro - {datetime.now().strftime('%d/%m/%Y %H:%M')}", title_style))
        elements.append(Spacer(1, 12))

        # 1. Resumo Executivo
        elements.append(Paragraph("Resumo Executivo", styles['Heading2']))
        summary_data = [
            ["Métrica", "Valor"],
            ["Saldo Atual", f"{summary.get('balance', 0):,.2f} Kz"],
            ["Total Receitas", f"{summary.get('income', 0):,.2f} Kz"],
            ["Total Despesas", f"{summary.get('expenses', 0):,.2f} Kz"],
        ]
        
        sm_table = Table(summary_data, colWidths=[150, 150])
        sm_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#334155")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(sm_table)
        elements.append(Spacer(1, 20))

        # 2. Tabela de Transações
        elements.append(Paragraph("Detalhes das Transações", styles['Heading2']))
        
        trans_data = [["Data", "Descrição", "Categoria", "Tipo", "Valor"]]
        for t in transactions[:100]: # Limitar a 100 para o PDF não ficar gigante
            trans_data.append([
                t.get('Data', 'N/A'),
                t.get('Descricao', 'Sem Descrição'),
                t.get('Categoria_nome', 'N/A'),
                t.get('Tipo', 'N/A'),
                f"{float(t.get('Valor', 0)):,.2f} Kz"
            ])

        t_table = Table(trans_data, colWidths=[80, 180, 100, 60, 100])
        t_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#475569")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white])
        ]))
        elements.append(t_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_excel_report(transactions):
        """
        Gera um relatório Excel (.xlsx).
        Retorna: BytesIO object contendo o Excel.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Transações"

        # Cabeçalho
        headers = ["Data", "Descrição", "Categoria", "Tipo", "Valor (Kz)"]
        ws.append(headers)

        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for t in transactions:
            ws.append([
                t.get('Data'),
                t.get('Descricao'),
                t.get('Categoria_nome'),
                t.get('Tipo'),
                float(t.get('Valor', 0))
            ])

        # Ajustar largura das colunas
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
